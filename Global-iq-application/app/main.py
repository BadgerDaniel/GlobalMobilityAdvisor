# app/main.py

# --- Imports ---
import chainlit as cl
from openai import AsyncOpenAI
import os
from dotenv import load_dotenv
import fitz  # PyMuPDF for PDFs
import docx  # python-docx for DOCX
import openpyxl # for XLSX
import csv      # built-in for CSV
import json     # built-in for JSON
import io       # built-in for file handling
from enhanced_agent_router import EnhancedAgentRouter
from input_collector import InputCollector
from typing import Optional
import hashlib
import httpx  # For MCP server communication
# Data persistence imports
from chainlit.data.sql_alchemy import SQLAlchemyDataLayer

# --- Load Environment Variables ---
load_dotenv()

# --- User Database (In production, use a proper database) ---
USERS_DB = {
    "admin": {
        "password_hash": hashlib.sha256("admin123".encode()).hexdigest(),
        "role": "admin",
        "name": "Administrator",
        "email": "admin@globaliq.com"
    },
    "hr_manager": {
        "password_hash": hashlib.sha256("hr2024".encode()).hexdigest(),
        "role": "hr_manager",
        "name": "HR Manager",
        "email": "hr@globaliq.com"
    },
    "employee": {
        "password_hash": hashlib.sha256("employee123".encode()).hexdigest(),
        "role": "employee",
        "name": "Employee User",
        "email": "employee@globaliq.com"
    },
    "demo": {
        "password_hash": hashlib.sha256("demo".encode()).hexdigest(),
        "role": "demo",
        "name": "Demo User",
        "email": "demo@globaliq.com"
    }
}

# --- Authentication Callback ---
@cl.password_auth_callback
def auth_callback(username: str, password: str) -> Optional[cl.User]:
    """
    Authenticate users with username and password.
    Returns a cl.User object if authentication is successful, None otherwise.
    """
    # Hash the provided password
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    # Check if user exists and password matches
    if username in USERS_DB and USERS_DB[username]["password_hash"] == password_hash:
        user_data = USERS_DB[username]
        return cl.User(
            identifier=username,
            metadata={
                "role": user_data["role"],
                "name": user_data["name"],
                "email": user_data["email"],
                "provider": "credentials"
            }
        )
    return None

# --- Initialize OpenAI Client ---
client = AsyncOpenAI(api_key=os.getenv("OPENAI_API_KEY"))
cl.instrument_openai()

# --- Initialize Enhanced Agent Router ---
router = EnhancedAgentRouter(api_key=os.getenv("OPENAI_API_KEY"))

# --- Initialize Input Collector ---
input_collector = InputCollector(openai_client=client)

# --- MCP Client Helper ---
class MCPClient:
    """Helper class for communicating with MCP servers"""

    def __init__(self):
        self.compensation_url = os.getenv("COMPENSATION_SERVER_URL", "http://localhost:8081")
        self.policy_url = os.getenv("POLICY_SERVER_URL", "http://localhost:8082")
        self.timeout = 30.0  # 30 second timeout

    async def predict_compensation(self, collected_data: dict) -> dict:
        """
        Call compensation MCP server

        Args:
            collected_data: Dictionary with keys like "Origin Location", "Destination Location", etc.

        Returns:
            Structured prediction response from MCP server
        """
        try:
            # Parse salary (handle formats like "100,000 USD" or "100k USD")
            salary_str = collected_data.get("Current Compensation", "0")
            salary = self._parse_salary(salary_str)

            # Build request payload
            payload = {
                "origin_location": collected_data.get("Origin Location", ""),
                "destination_location": collected_data.get("Destination Location", ""),
                "current_salary": salary,
                "currency": self._extract_currency(salary_str),
                "assignment_duration": collected_data.get("Assignment Duration", "12 months"),
                "job_level": collected_data.get("Job Level/Title", "Manager"),
                "family_size": int(collected_data.get("Family Size", "1")),
                "housing_preference": collected_data.get("Housing Preference", "Company-provided")
            }

            async with httpx.AsyncClient() as http_client:
                response = await http_client.post(
                    f"{self.compensation_url}/predict",
                    json=payload,
                    timeout=self.timeout
                )
                response.raise_for_status()
                return response.json()

        except Exception as e:
            print(f"MCP compensation prediction error: {e}")
            raise

    async def analyze_policy(self, collected_data: dict) -> dict:
        """
        Call policy MCP server

        Args:
            collected_data: Dictionary with keys like "Origin Country", "Destination Country", etc.

        Returns:
            Structured policy analysis from MCP server
        """
        try:
            # Build request payload
            payload = {
                "origin_country": collected_data.get("Origin Country", ""),
                "destination_country": collected_data.get("Destination Country", ""),
                "assignment_type": collected_data.get("Assignment Type", "Long-term"),
                "duration": collected_data.get("Assignment Duration", "12 months"),
                "job_title": collected_data.get("Job Title", "Manager")
            }

            async with httpx.AsyncClient() as http_client:
                response = await http_client.post(
                    f"{self.policy_url}/analyze",
                    json=payload,
                    timeout=self.timeout
                )
                response.raise_for_status()
                return response.json()

        except Exception as e:
            print(f"MCP policy analysis error: {e}")
            raise

    def _parse_salary(self, salary_str: str) -> float:
        """Parse salary string to float"""
        try:
            # Remove common formatting
            cleaned = salary_str.replace(",", "").replace("$", "").strip()

            # Handle "100k" format
            if "k" in cleaned.lower():
                cleaned = cleaned.lower().replace("k", "")
                return float(cleaned) * 1000

            # Extract just the number part (before currency)
            parts = cleaned.split()
            return float(parts[0])
        except:
            return 0.0

    def _extract_currency(self, salary_str: str) -> str:
        """Extract currency from salary string"""
        salary_str_upper = salary_str.upper()
        if "USD" in salary_str_upper or "$" in salary_str:
            return "USD"
        elif "EUR" in salary_str_upper or "€" in salary_str:
            return "EUR"
        elif "GBP" in salary_str_upper or "£" in salary_str:
            return "GBP"
        else:
            return "USD"  # Default

# Initialize MCP Client
mcp_client = MCPClient()

# --- Data Persistence Layer Configuration ---
# Temporarily disabled due to user/thread integration issues
# @cl.data_layer
# def get_data_layer():
#     """
#     Configure SQLAlchemy data layer for chat history and multi-chat functionality.
#     Uses SQLite for local storage - no external dependencies required.
#     """
#     # Create SQLite database in the app directory
#     db_path = os.path.join(os.path.dirname(__file__), "chat_history.db")
#     conninfo = f"sqlite+aiosqlite:///{db_path}"
#     
#     return SQLAlchemyDataLayer(conninfo=conninfo)

# --- System Prompt for LLM ---
def get_system_prompt(user_name: str = "Guest", user_role: str = "guest") -> str:
    """Generate a dynamic system prompt based on user context."""
    base_prompt = f"""You are the Global IQ Mobility Advisor, an expert HR assistant helping {user_name} (Role: {user_role.replace('_', ' ').title()}).
Your goal is to help plan employee relocations based on user requests.

User Role Context:
"""
    
    if user_role == 'admin':
        base_prompt += "- You're assisting an Administrator with full system access\n"
        base_prompt += "- Provide detailed technical information and system insights\n"
        base_prompt += "- Include administrative recommendations and policy suggestions\n"
    elif user_role == 'hr_manager':
        base_prompt += "- You're assisting an HR Manager with policy and employee data access\n"
        base_prompt += "- Focus on policy compliance, employee welfare, and cost management\n"
        base_prompt += "- Provide detailed compensation and policy analysis\n"
    elif user_role == 'employee':
        base_prompt += "- You're assisting an Employee with personal relocation needs\n"
        base_prompt += "- Focus on practical guidance and personal impact\n"
        base_prompt += "- Explain policies in easy-to-understand terms\n"
    elif user_role == 'demo':
        base_prompt += "- You're assisting a Demo User exploring the system\n"
        base_prompt += "- Provide comprehensive examples and explanations\n"
        base_prompt += "- Highlight system capabilities and features\n"
    
    base_prompt += "\nInstructions:\n"
    base_prompt += "- First, check if context from attached documents (like HR policies, guides, data files) is provided\n"
    base_prompt += "- If documents are provided, prioritize that information and clearly state your answer is based on the provided document(s)\n"
    base_prompt += "- If no specific document context is provided, state that you lack specific data but can answer general questions\n"
    base_prompt += "- Maintain conversation history context and be professional, concise, and helpful\n"
    base_prompt += f"- Always address the user as {user_name} when appropriate"
    
    return base_prompt

# --- File Processing Functions ---

def process_pdf(file_path: str) -> str:
    """Extracts text from a PDF file using PyMuPDF."""
    text = ""
    try:
        doc = fitz.open(file_path)
        for page in doc:
            page_text = page.get_text("text")
            if page_text:
                text += page_text + "\n\n"
        doc.close()
    except Exception as e:
        print(f"Error processing PDF {file_path}: {e}")
        raise # Re-raise exception to be caught in handler
    return text.strip()

def process_docx(file_path: str) -> str:
    """Extracts text from a DOCX file using python-docx."""
    text = ""
    try:
        document = docx.Document(file_path)
        for para in document.paragraphs:
            text += para.text + "\n\n"
    except Exception as e:
        print(f"Error processing DOCX {file_path}: {e}")
        raise
    return text.strip()

def process_txt(file_path: str) -> str:
    """Reads text from a TXT file."""
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()
    except Exception as e:
        print(f"Error processing TXT {file_path}: {e}")
        raise
    return text.strip()

def process_csv(file_path: str) -> str:
    """Reads content from a CSV file and formats as text."""
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline='') as f:
            reader = csv.reader(f)
            for row in reader:
                # Join cells with a delimiter (e.g., comma and space)
                text += ", ".join(cell.strip() for cell in row) + "\n"
    except Exception as e:
        print(f"Error processing CSV {file_path}: {e}")
        raise
    return text.strip()

def process_json(file_path: str) -> str:
    """Reads content from a JSON file and formats as text."""
    text = ""
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            data = json.load(f)
            # Pretty-print the JSON data as text
            text = json.dumps(data, indent=2)
    except Exception as e:
        print(f"Error processing JSON {file_path}: {e}")
        raise
    return text.strip()

def process_xlsx(file_path: str) -> str:
    """Extracts text from an XLSX file using openpyxl."""
    text = ""
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            text += f"--- Sheet: {sheet_name} ---\n"
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows():
                row_texts = []
                for cell in row:
                    if cell.value is not None:
                        row_texts.append(str(cell.value).strip())
                if row_texts: # Only add non-empty rows
                    text += ", ".join(row_texts) + "\n"
            text += "\n" # Add space between sheets
    except Exception as e:
        print(f"Error processing XLSX {file_path}: {e}")
        raise
    return text.strip()

# --- File Handler Dispatch Dictionary ---
# Maps MIME types (or extensions as fallback) to processing functions
FILE_HANDLERS = {
    "application/pdf": process_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": process_docx,
    ".docx": process_docx, # Fallback for extension
    "text/plain": process_txt,
    ".txt": process_txt,
    "text/csv": process_csv,
    ".csv": process_csv,
    "application/json": process_json,
    ".json": process_json,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": process_xlsx,
    ".xlsx": process_xlsx,
}

# --- Calculation Functions (Using MCP) ---
async def _run_compensation_calculation(collected_data: dict, extracted_texts: list) -> str:
    """Calculate compensation using MCP server."""
    try:
        # Call MCP server for prediction
        mcp_response = await mcp_client.predict_compensation(collected_data)

        # Check if successful
        if mcp_response.get("status") != "success":
            raise Exception(mcp_response.get("error", "Unknown error"))

        # Extract prediction data
        predictions = mcp_response.get("predictions", {})
        breakdown = mcp_response.get("breakdown", {})
        confidence = mcp_response.get("confidence_scores", {})
        recommendations = mcp_response.get("recommendations", [])
        methodology = mcp_response.get("methodology", {})

        # Format response in a professional manner
        result = f"""💰 **Compensation Calculation Results**

**Summary:**
• **Origin:** {collected_data.get("Origin Location", "N/A")}
• **Destination:** {collected_data.get("Destination Location", "N/A")}
• **Current Salary:** {collected_data.get("Current Compensation", "N/A")}

**Compensation Package Breakdown:**

**Base Salary:** {predictions.get('currency', 'USD')} {predictions.get('base_salary', 0):,.2f}
**COLA Ratio:** {predictions.get('cola_ratio', 1.0):.2f}x
**Adjusted Salary:** {predictions.get('currency', 'USD')} {predictions.get('adjusted_salary', 0):,.2f}
**Housing Allowance:** {predictions.get('currency', 'USD')} {predictions.get('housing_allowance', 0):,.2f}
**Hardship Pay:** {predictions.get('currency', 'USD')} {predictions.get('hardship_pay', 0):,.2f}

**═══════════════════════════════════**
**Total Package:** {predictions.get('currency', 'USD')} {predictions.get('total_package', 0):,.2f}
**═══════════════════════════════════**

**Confidence Scores:**
• COLA: {confidence.get('cola', 0.0):.0%}
• Housing: {confidence.get('housing', 0.0):.0%}
• Overall: {confidence.get('overall', 0.0):.0%}

**Recommendations:**
"""
        # Add recommendations
        for i, rec in enumerate(recommendations, 1):
            result += f"{i}. {rec}\n"

        # Add methodology info
        result += f"\n**Methodology:** {methodology.get('model_type', 'N/A')} (v{methodology.get('version', 'N/A')})"

        return result

    except Exception as e:
        error_msg = f"Sorry, I encountered an error with the MCP compensation service: {str(e)}"
        print(error_msg)
        # Fallback: return a simple error message
        return f"❌ **Error**: {error_msg}\n\nPlease ensure the MCP servers are running."

async def _run_policy_analysis(collected_data: dict, extracted_texts: list) -> str:
    """Analyze policy using MCP server."""
    try:
        # Call MCP server for analysis
        mcp_response = await mcp_client.analyze_policy(collected_data)

        # Check if successful
        if mcp_response.get("status") != "success":
            raise Exception(mcp_response.get("error", "Unknown error"))

        # Extract analysis data
        analysis = mcp_response.get("analysis", {})
        recommendations = mcp_response.get("recommendations", [])
        confidence = mcp_response.get("confidence", 0.0)
        metadata = mcp_response.get("metadata", {})

        # Extract visa requirements
        visa_req = analysis.get("visa_requirements", {})
        eligibility = analysis.get("eligibility", {})
        compliance = analysis.get("compliance", {})
        timeline = analysis.get("timeline", {})
        documentation = analysis.get("documentation", [])

        # Format response
        result = f"""📋 **Policy Analysis Results**

**Summary:**
• **Origin:** {collected_data.get("Origin Country", "N/A")}
• **Destination:** {collected_data.get("Destination Country", "N/A")}
• **Assignment Type:** {collected_data.get("Assignment Type", "N/A")}
• **Duration:** {collected_data.get("Assignment Duration", "N/A")}

**Visa Requirements:**
• **Type:** {visa_req.get('visa_type', 'N/A')}
• **Processing Time:** {visa_req.get('processing_time', 'N/A')}
• **Cost:** {visa_req.get('cost', 'N/A')}
• **Requirements:**
"""
        for req in visa_req.get('requirements', []):
            result += f"  - {req}\n"

        result += f"""
**Eligibility:**
• **Meets Requirements:** {'Yes' if eligibility.get('meets_requirements', False) else 'No'}
"""
        if eligibility.get('concerns'):
            result += "• **Concerns:**\n"
            for concern in eligibility.get('concerns', []):
                result += f"  - {concern}\n"

        result += f"""
**Timeline:**
• **Visa Application:** {timeline.get('visa_application', 'N/A')}
• **Visa Approval:** {timeline.get('visa_approval', 'N/A')}
• **Relocation Prep:** {timeline.get('relocation_prep', 'N/A')}
• **Target Start Date:** {timeline.get('start_date', 'N/A')}

**Required Documentation:**
"""
        for doc in documentation:
            result += f"• {doc}\n"

        result += f"""
**Compliance Considerations:**
**Origin Country Requirements:**
"""
        for req in compliance.get('origin_requirements', []):
            result += f"• {req}\n"

        result += "\n**Destination Country Requirements:**\n"
        for req in compliance.get('destination_requirements', []):
            result += f"• {req}\n"

        result += "\n**Recommendations:**\n"
        for i, rec in enumerate(recommendations, 1):
            result += f"{i}. {rec}\n"

        result += f"\n**Confidence:** {confidence:.0%}"
        result += f"\n**Data Source:** {metadata.get('model_type', 'N/A')} (Updated: {metadata.get('last_updated', 'N/A')})"

        return result

    except Exception as e:
        error_msg = f"Sorry, I encountered an error with the MCP policy service: {str(e)}"
        print(error_msg)
        return f"❌ **Error**: {error_msg}\n\nPlease ensure the MCP servers are running."

# --- Chainlit Event Handlers ---

@cl.on_chat_start
async def start_chat():
    """Initializes session history and welcomes authenticated user."""
    cl.user_session.set("history", [])
    
    # Get the authenticated user
    user = cl.user_session.get("user")
    
    if user:
        # Welcome message with user information
        welcome_msg = f"🌍 **Welcome to Global IQ Mobility Advisor, {user.metadata['name']}!**\n\n"
        welcome_msg += f"**User Role:** {user.metadata['role'].replace('_', ' ').title()}\n"
        welcome_msg += f"**Email:** {user.metadata['email']}\n\n"
        
        # Role-specific welcome messages
        if user.metadata['role'] == 'admin':
            welcome_msg += "🔧 **Admin Access:** You have full access to all features including user management and system analytics.\n\n"
        elif user.metadata['role'] == 'hr_manager':
            welcome_msg += "👥 **HR Manager Access:** You can access employee data, policy information, and compensation calculations.\n\n"
        elif user.metadata['role'] == 'employee':
            welcome_msg += "👤 **Employee Access:** You can get assistance with relocation policies and compensation inquiries.\n\n"
        elif user.metadata['role'] == 'demo':
            welcome_msg += "🎯 **Demo Access:** Explore the Global IQ features with sample data and scenarios.\n\n"
        
        welcome_msg += "I'm here to help you with:\n"
        welcome_msg += "• 📋 **Policy Information** - HR policies, visa requirements, relocation guidelines\n"
        welcome_msg += "• 💰 **Compensation Analysis** - Salary adjustments, cost of living calculations\n"
        welcome_msg += "• 📄 **Document Analysis** - Upload and analyze HR documents, policies, and data files\n\n"
        welcome_msg += "**How can I assist you today?**"
        
        await cl.Message(content=welcome_msg).send()
    else:
        # Fallback if user session is not properly set
        await cl.Message(content="🌍 Welcome to Global IQ Mobility Advisor! How can I help you today?").send()
    
    print(f"Chat session started for user: {user.identifier if user else 'Unknown'}")

@cl.on_message
async def handle_message(msg: cl.Message):
    """Handles messages, processes files, calls LLM."""
    # Get authenticated user
    user = cl.user_session.get("user")
    user_role = user.metadata['role'] if user else 'guest'
    user_name = user.metadata['name'] if user else 'Guest'
    
    # Handle special admin commands
    if msg.content and user_role == 'admin':
        if msg.content.lower() == '/users':
            user_list = "👥 **Current Users:**\n\n"
            for username, data in USERS_DB.items():
                user_list += f"• **{username}** - {data['name']} ({data['role']}) - {data['email']}\n"
            await cl.Message(content=user_list).send()
            return
        elif msg.content.lower() == '/help':
            help_msg = "🔧 **Admin Commands:**\n\n"
            help_msg += "• `/users` - List all registered users\n"
            help_msg += "• `/help` - Show this help message\n"
            help_msg += "• `/history` - View current session chat history\n\n"
            help_msg += "**Multi-Chat Features:**\n"
            help_msg += "• Chat history is now automatically saved\n"
            help_msg += "• Users can resume previous conversations\n"
            help_msg += "• All conversations are persistent across sessions\n\n"
            help_msg += "**Regular features are also available for policy and compensation analysis.**"
            await cl.Message(content=help_msg).send()
            return
        elif msg.content.lower() == '/history':
            history = cl.user_session.get("history", [])
            if history:
                history_msg = "📜 **Current Session Chat History:**\n\n"
                for i, item in enumerate(history[-10:], 1):  # Show last 10 messages
                    role_emoji = "👤" if item['role'] == 'user' else "🤖"
                    content_preview = item['content'][:100] + "..." if len(item['content']) > 100 else item['content']
                    history_msg += f"{i}. {role_emoji} **{item['role'].capitalize()}:** {content_preview}\n"
                if len(history) > 10:
                    history_msg += f"\n*Showing last 10 of {len(history)} total messages*"
            else:
                history_msg = "📜 **Chat History:** No messages in current session yet."
            await cl.Message(content=history_msg).send()
            return
    
    history = cl.user_session.get("history", []) # Default to empty list if not found
    if msg.content:
        history.append({"role": "user", "content": msg.content})

    # --- Process Attached Files using Dispatch Dict ---
    extracted_texts = []
    unsupported_files = []
    if msg.elements:
        await cl.Message(content="Processing attached file(s)...").send()
        for element in msg.elements:
            handler = FILE_HANDLERS.get(element.mime) # Try MIME type first
            if not handler and element.name: # If no MIME handler, try extension
                ext = os.path.splitext(element.name.lower())[1]
                handler = FILE_HANDLERS.get(ext)

            if handler:
                try:
                    print(f"--- DEBUG: Processing {element.name} using {handler.__name__} ---")
                    # We might need to wrap synchronous file processing if it's slow
                    # using await cl.make_async or asyncio.to_thread
                    # For now, calling directly for simplicity:
                    file_content = handler(element.path)

                    if file_content:
                        extracted_texts.append({
                            "name": element.name,
                            "content": file_content
                        })
                        await cl.Message(content=f"Successfully processed text from `{element.name}`.").send()
                    else:
                        await cl.Message(content=f"Could not extract meaningful text from `{element.name}`.").send()
                except Exception as e:
                    print(f"Error processing file {element.name}: {e}")
                    await cl.Message(content=f"Sorry, encountered an error processing file `{element.name}`.").send()
            else:
                # Handle unsupported file types
                unsupported_files.append(element.name)
                print(f"--- DEBUG: Unsupported file type: {element.name} ({element.mime}) ---")

        # Send a single message about unsupported files at the end
        if unsupported_files:
            unsupported_list = ", ".join([f"`{f}`" for f in unsupported_files])
            supported_types = ".pdf, .docx, .xlsx, .csv, .json, .txt" # Update as needed
            await cl.Message(content=f"Note: Could not process the following file(s) as the type is not supported: {unsupported_list}.\nWe currently support: {supported_types}").send()

    # --- Prepare Prompt for LLM ---
    # Generate dynamic system prompt based on user context
    system_prompt = get_system_prompt(user_name, user_role)
    prompt_messages = [{"role": "system", "content": system_prompt}]
    if extracted_texts:
        file_context = "Context from attached document(s):\n\n"
        for item in extracted_texts:
            max_len = 3000 # Max length per file context
            truncated_content = item['content'][:max_len]
            if len(item['content']) > max_len:
                truncated_content += "..."
            file_context += f"--- Document: {item['name']} ---\n{truncated_content}\n\n"
        prompt_messages.append({"role": "system", "content": file_context.strip()})
        print("--- DEBUG: Added file context to prompt ---")
    prompt_messages.extend(history)

    # --- Enhanced Routing with Input Collection ---
    try:
        # Get user session data
        user_session = cl.user_session.get("user_data", {})
        user_query = msg.content or "General inquiry"
        
        # Check if we're in the middle of input collection or awaiting responses
        in_compensation_collection = input_collector.is_collection_in_progress("compensation", user_session)
        in_policy_collection = input_collector.is_collection_in_progress("policy", user_session)
        awaiting_both_choice = user_session.get("awaiting_both_choice", False)
        awaiting_compensation_confirmation = user_session.get("awaiting_compensation_confirmation", False)
        awaiting_policy_confirmation = user_session.get("awaiting_policy_confirmation", False)
        
        if in_compensation_collection:
            # Continue compensation collection
            response, updated_session, is_completed = input_collector.process_answer(
                "compensation", user_query, user_session
            )
            cl.user_session.set("user_data", updated_session)
            
            await cl.Message(content=response).send()
            
            if is_completed:
                # Start compensation calculation
                collected_data = input_collector.get_collected_data("compensation", updated_session)
                calc_response = await _run_compensation_calculation(collected_data, extracted_texts)
                await cl.Message(content=calc_response).send()
            
            return
            
        elif in_policy_collection:
            # Continue policy collection
            response, updated_session, is_completed = input_collector.process_answer(
                "policy", user_query, user_session
            )
            cl.user_session.set("user_data", updated_session)
            
            await cl.Message(content=response).send()
            
            if is_completed:
                # Start policy analysis
                collected_data = input_collector.get_collected_data("policy", updated_session)
                policy_response = await _run_policy_analysis(collected_data, extracted_texts)
                await cl.Message(content=policy_response).send()
            
            return
        
        # Handle confirmation responses
        if awaiting_compensation_confirmation:
            user_response = user_query.lower().strip()
            user_session["awaiting_compensation_confirmation"] = False
            
            if any(word in user_response for word in ['yes', 'start', 'ok', 'sure', 'proceed']):
                # Start compensation collection
                question, updated_session = input_collector.start_collection("compensation", user_session)
                cl.user_session.set("user_data", updated_session)
                await cl.Message(content=f"💰 **Starting Compensation Analysis**\n\n{question}").send()
                return
            elif any(word in user_response for word in ['no', 'not', 'different', 'wrong']):
                # Show general help instead
                help_message = input_collector.get_general_help_message()
                await cl.Message(content=help_message).send()
                cl.user_session.set("user_data", user_session)
                return
            else:
                # Unclear response, ask again
                conf_message = input_collector.get_confirmation_message("compensation")
                await cl.Message(content=f"I didn't understand. {conf_message}").send()
                user_session["awaiting_compensation_confirmation"] = True
                cl.user_session.set("user_data", user_session)
                return
        
        if awaiting_policy_confirmation:
            user_response = user_query.lower().strip()
            user_session["awaiting_policy_confirmation"] = False
            
            if any(word in user_response for word in ['yes', 'start', 'ok', 'sure', 'proceed']):
                # Start policy collection
                question, updated_session = input_collector.start_collection("policy", user_session)
                cl.user_session.set("user_data", updated_session)
                await cl.Message(content=f"📋 **Starting Policy Analysis**\n\n{question}").send()
                return
            elif any(word in user_response for word in ['no', 'not', 'different', 'wrong']):
                # Show general help instead
                help_message = input_collector.get_general_help_message()
                await cl.Message(content=help_message).send()
                cl.user_session.set("user_data", user_session)
                return
            else:
                # Unclear response, ask again
                conf_message = input_collector.get_confirmation_message("policy")
                await cl.Message(content=f"I didn't understand. {conf_message}").send()
                user_session["awaiting_policy_confirmation"] = True
                cl.user_session.set("user_data", user_session)
                return
        
        # Handle "both choice" responses
        if awaiting_both_choice:
            user_choice = user_query.lower().strip()
            user_session["awaiting_both_choice"] = False
            
            if "policy" in user_choice or user_choice == "1":
                # Start policy collection
                question, updated_session = input_collector.start_collection("policy", user_session)
                cl.user_session.set("user_data", updated_session)
                await cl.Message(content=f"📋 **Starting with Policy Analysis**\n\n{question}").send()
                return
            elif "compensation" in user_choice or user_choice == "2":
                # Start compensation collection
                question, updated_session = input_collector.start_collection("compensation", user_session)
                cl.user_session.set("user_data", updated_session)
                await cl.Message(content=f"💰 **Starting with Compensation Calculation**\n\n{question}").send()
                return
            else:
                # Invalid choice, show options again
                choice_message = input_collector.get_both_choice_message()
                await cl.Message(content=f"I didn't understand your choice. {choice_message}").send()
                user_session["awaiting_both_choice"] = True
                cl.user_session.set("user_data", user_session)
                return
        
        # New query - route it
        routing_msg = cl.Message(content="🔄 Analyzing your query and routing to the appropriate specialist...")
        await routing_msg.send()
        
        routing_result = router.route_query(user_query)
        route_name = routing_result["destination"]
        routing_method = routing_result.get("routing_method", "unknown")
        
        # Get enhanced display info
        display_info = router.get_route_display_info(route_name)
        routing_update = f"{display_info['emoji']} **Routed to: {display_info['title']}** ({routing_method} routing)\n\n*{display_info['description']}*"
        routing_msg.content = routing_update
        await routing_msg.update()
        
        # Handle different route types
        if route_name == "compensation":
            # Show confirmation message before starting collection
            conf_message = input_collector.get_confirmation_message("compensation")
            await cl.Message(content=conf_message).send()
            user_session["awaiting_compensation_confirmation"] = True
            cl.user_session.set("user_data", user_session)
            
        elif route_name == "policy":
            # Show confirmation message before starting collection
            conf_message = input_collector.get_confirmation_message("policy")
            await cl.Message(content=conf_message).send()
            user_session["awaiting_policy_confirmation"] = True
            cl.user_session.set("user_data", user_session)
            
        elif route_name == "both_policy_and_compensation":
            # Show choice message
            choice_message = input_collector.get_both_choice_message()
            await cl.Message(content=choice_message).send()
            # Set a flag to handle the next response as a choice
            user_session["awaiting_both_choice"] = True
            cl.user_session.set("user_data", user_session)
            
        elif route_name == "guidance_fallback":
            # Only show intro message if not shown before in this session
            if not user_session.get("intro_shown", False):
                intro_message = input_collector.get_intro_message()
                await cl.Message(content=intro_message).send()
                user_session["intro_shown"] = True
                cl.user_session.set("user_data", user_session)
            else:
                # If intro already shown, try to be more helpful
                helpful_msg = "I'm here to help with global mobility questions. You can ask about:\n\n"
                helpful_msg += "💰 **Compensation** - 'Calculate salary for moving to [city]'\n"
                helpful_msg += "📋 **Policy** - 'What are the rules for [assignment type]?'\n"
                helpful_msg += "🎯 **Strategic** - 'What's the best way to relocate [role]?'\n\n"
                helpful_msg += "What specific mobility scenario can I help you with?"
                await cl.Message(content=helpful_msg).send()
            
        else:
            # Fallback to original routing
            context_info = ""
            if extracted_texts:
                context_info = "\n\nContext from uploaded documents:\n"
                for item in extracted_texts:
                    max_len = 1500
                    truncated_content = item['content'][:max_len]
                    if len(item['content']) > max_len:
                        truncated_content += "..."
                    context_info += f"\n--- {item['name']} ---\n{truncated_content}\n"
            
            enhanced_input = user_query + context_info
            routing_result["next_inputs"]["input"] = enhanced_input
            
            response = router.get_route_response(
                routing_result["destination"], 
                routing_result["next_inputs"]
            )
            
            await cl.Message(content=response).send()
            
    except Exception as e:
        error_message = f"Sorry, I encountered an error processing your request: {e}"
        await cl.Message(content=error_message).send()
        print(f"Error during OpenAI API call/streaming: {e}")